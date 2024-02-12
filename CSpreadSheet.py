#!/usr/bin/python
import csv
import os
import sys

# Module openpyxl needs to be imported separately
try:
    import openpyxl
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print("Couldn't import openpyxl. Install with 'pip install openpyxl'.")
    sys.exit(1)


class CSpreadSheet:
    def __init__(self, path):
        self.path = path
        self.dirname = os.path.dirname(path)
        self.filename = os.path.basename(os.path.splitext(path)[0])
        self.extension = os.path.splitext(path)[1]
        self.workbook = None

        # Load the file
        if self.extension == ".xlsx":
            self.workbook = load_workbook(filename=self.path)
        elif self.extension == ".csv":
            # We only have one sheet in a csv file
            title = 'active'
            # For uniformity, write into a xlxs workbook
            self.workbook = openpyxl.Workbook()
            self.workbook.create_sheet(title)
            with open(self.path, 'r') as f:
                reader = csv.reader(f, delimiter='\t')
                for r, row in enumerate(reader):
                    for c, value in enumerate(row):
                        self.setValue(title, r+1, c+1, value)
        else:
            print(f"! Didn't recognise extension '{self.extension}'")

    def validate(self, sheetName, headers):
        # Check if workbook and sheet have been created
        if sheetName not in self.workbook:
            print(f"! Couldn't find sheet with name '{sheetName}'.")
            return 1

        sheet = self.workbook[sheetName]

        # Check if it contains at least a header row and one row of data
        if sheet.max_row <= 2:
            print(f"! Sheet '{sheetName}' doesn't contain enough rows of data")
            return 1

        # Check the columns
        if sheet.max_column != len(headers):
            print(f"! Number of columns in sheet '{sheetName}' should be {len(headers)}, got {sheet.max_column}")
            return 1
        for index, header in enumerate(sheet[1]):
            if header.value != headers[index]:
                print(f"! Unexpected header in sheet '{sheetName}': got '{header.value}', expected '{headers[index]}'")
                return 1

        # Check the rows
        lastRowValid = False
        emptyLines = 0
        # Spreadsheets sometimes have last rows empty after real data. Find out
        # where real data starts
        while not lastRowValid and sheet.max_row > 2:
            if (sheet.cell(row=sheet.max_row, column=1).value is None and
                sheet.cell(row=sheet.max_row, column=2).value is None):
                sheet.delete_rows(sheet.max_row)
                emptyLines += 1
                continue
            lastRowValid = True
        if emptyLines:
            print(f"      ! Deleted last {emptyLines} rows as empty in sheet '{sheetName}'")

        if sheet.max_row < 2:
            print(f"      ! No data in rows for sheet '{sheetName}'")
            return 1

        return 0

    def getValue(self, sheetName, rowIndex, columnIndex):
        sheet = self.workbook[sheetName]
        return sheet.cell(row=rowIndex, column=columnIndex).value

    def getColumn(self, sheetName, columnIndex):
        values = []
        sheet = self.workbook[sheetName]
        for currentRow in range(1, sheet.max_row):
            value = sheet.cell(row=currentRow, column=columnIndex).value
            values.append(value)
        return values

    def setValue(self, sheetIndex, rowIndex, columnIndex, newValue):
        sheet = self.workbook[sheetIndex]
        return sheet.cell(row=rowIndex, column=columnIndex, value=newValue)

    def save(self, newPath=None):
        if not newPath:
            newPath = self.path
        if self.extension == ".xlsx":
            try:
                self.workbook.save(filename=newPath)
                return 0
            except PermissionError:
                print(f"! Couldn't write to '{newPath}'. Still open?")
                return 1
        elif self.extension == ".csv":
            with open(newPath, 'w', newline='') as file:
                writer = csv.writer(file, delimiter='\t')
                for row in self.workbook['active'].rows:
                    writer.writerow([cell.value for cell in row])
            return 0
        else:
            print(f"! Didn't recognise extension {'self.extension'}")
            return 1
